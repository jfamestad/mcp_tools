from getopt import short_has_arg

import openpyxl

from mcp.server.fastmcp import FastMCP

# Pass lifespan to server
mcp = FastMCP("excel workbook actions")

class Workbook:
	"""
	Wrapper for openpyxl

	Parameters
	----------
	file_path : str
		Path to the excel file
	"""
	def __init__(self, file_path):
		"""
		:param file_path:
		"""
		self.file_path = file_path
		self.get_workbook_data(file_path)

	def get_workbook_data(self, file_path):
		"""
		Get the data from the workbook
		:param file_path:
		:return: openpyxl Workbook object
		"""
		self._data = openpyxl.load_workbook(file_path)

	def list_sheets(self) -> list[str]:
		"""
		List the titles of the sheets in the workbook
		:return: list of sheet names (str)

		Example
		-------
		wb.list_sheets()
		['MainSheet', 'AnotherSheet']
		"""
		return self._data.sheetnames

	def get_sheet(self, sheet_name) -> openpyxl.worksheet.worksheet.Worksheet:
		"""
		Get a sheet from the workbook
		:param sheet_name:
		:return: openpyxl Worksheet
		"""
		# use get_sheet_by_name(name)[source] instead
		# return [ it for it in self._data._sheets if it.title == sheet_name ][0]
		return self._data[sheet_name]

	def save(self):
		"""
		Save the workbook to disk
		:return:
		"""
		self._data.save(self.file_path)

	def read_row(self, sheet_name, row):
		"""
		Read a row from a sheet
		:param sheet_name:
		:param row:
		:return:
		"""
		sheet = self.get_sheet(sheet_name)
		return sheet[row]

	def read_rows(self, sheet_name):
		"""
		Read all rows from a sheet
		:param sheet_name:
		:return:
		"""
		sheet = self.get_sheet(sheet_name)
		rows = []
		for row in sheet:
			row_values = [it.value for it in row]
			rows.append(row_values)
		return rows

	def add_row(self, sheet_name, row, skip_save=False):
		"""
		Add a row to a sheet
		:param sheet_name:
		:param row:
		:param skip_save:
		:return:
		"""
		sheet = self.get_sheet(sheet_name)
		sheet.append(row)
		if not skip_save:
			self.save()

	def add_rows(self, sheet_name, rows):
		"""
		Add multiple rows to a sheet
		:param sheet_name:
		:param rows:
		:return:
		"""
		for row in rows:
			self.add_row(sheet_name, row, skip_save=True)
		self.save()

	def update_row(self, sheet_name, row_number, row, skip_save=False):
		"""
		Update a row
		:param sheet_name:
		:param row_number:
		:param row:
		:param skip_save:
		:return:
		"""
		sheet = self.get_sheet(sheet_name)
		for i, cell_data in enumerate(row):
			sheet.cell(row_number, i + 1).value = cell_data
		if not skip_save:
			self.save()

	def copy_worksheet(self, source_sheet_name, target_sheet_name):
		"""
		Copy a worksheet
		:param source_sheet_name:
		:param target_sheet_name:
		:return:
		"""
		new_sheet = self._data.copy_worksheet(self.get_sheet(source_sheet_name))
		new_sheet.title = target_sheet_name
		self.save()

	def delete_row(self, sheet_name, row_number):
		"""
		Delete a row
		:param sheet_name:
		:param row_number:
		:return:
		"""
		sheet = self.get_sheet(sheet_name)
		sheet.delete_rows(row_number)
		self.save()

	def replace_rows(self, first_row_index, sheet_name, rows, skip_save=False):
		current_row_index = first_row_index
		for row in rows:
			self.update_row(sheet_name, current_row_index, row)
			current_row_index += 1
		if not skip_save:
			self.save()






@mcp.tool('list_sheets_in_workbook')
def list_sheets_in_workbook(filepath: str):
	"""
	List the sheets in a workbook
	:param filepath:
	:return:
	"""
	wb = Workbook(filepath)
	return wb.list_sheets()

@mcp.tool('read_excel_sheet')
def read_excel_sheet(filepath: str, sheet_name: str):
	"""
	Read an excel sheet
	:param filepath:
	:param sheet_name:
	:return:
	"""
	wb = Workbook(filepath)
	return wb.read_rows(sheet_name)

@mcp.tool('add_rows_to_excel_sheet')
def add_rows_to_excel_sheet(filepath: str, sheet_name: str, rows: list):
	"""
	Add rows to an excel sheet
	:param filepath:
	:param sheet_name:
	:param rows:
	:return:
	"""
	wb = Workbook(filepath)
	wb.add_rows(sheet_name, rows)

@mcp.tool('append_row_to_excel_sheet')
def append_row_to_excel_sheet(filepath: str, sheet_name: str, row: list):
	"""
	Add a row to an excel sheet
	:param filepath:
	:param sheet_name:
	:param row:
	:return:
	"""
	wb = Workbook(filepath)
	wb.add_row(sheet_name, row)

@mcp.tool('update_row_in_excel_sheet')
def update_row_in_excel_sheet(filepath: str, sheet_name: str, row_number: int, row: list):
	"""
	Update a row in an excel sheet
	:param filepath:
	:param sheet_name:
	:param row_number:
	:param row:
	:return:
	"""
	wb = Workbook(filepath)
	wb.update_row(sheet_name, row_number, row)

@mcp.tool('copy_excel_sheet')
def copy_excel_sheet(filepath: str, source_sheet_name: str, target_sheet_name: str):
	"""
	Copy an excel sheet
	:param filepath:
	:param source_sheet_name:
	:param target_sheet_name:
	:return:
	"""
	wb = Workbook(filepath)
	wb.copy_worksheet(source_sheet_name, target_sheet_name)

@mcp.tool('delete-row-in-excel-sheet')
def delete_row_in_excel_sheet(filepath: str, sheet_name: str, row_number: int):
	"""
	Delete a row in an excel sheet
	:param filepath:
	:param sheet_name:
	:param row_number:
	:return:
	"""
	wb = Workbook(filepath)
	wb.delete_row(sheet_name, row_number)

@mcp.tool('replace-rows-in-excel-sheet')
def replace_rows_in_excel_sheet(filepath: str, sheet_name: str, first_row_index: int, rows: list):
	"""
	Replace rows in an excel sheet
	:param filepath:
	:param sheet_name:
	:param first_row_index:
	:param rows:
	:return:
	"""
	wb = Workbook(filepath)
	wb.replace_rows(first_row_index, sheet_name, rows)


@mcp.prompt()
def example_prompt() -> str:
    return f"Read data from {mcp.workbook.file_path}"

if __name__ == "__main__":
    mcp.run()